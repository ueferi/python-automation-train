import openpyxl as xl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

book = xl.Workbook()
sheet = book.active

# 横幅を指定
sheet.column_dimensions["B"].width = 40
# 高さを指定
sheet.row_dimensions[2].height = 40

cell = sheet["B2"]
cell.value = "喜びにあふれた心は良い薬"

# テキスト配置の指定
# 水平位置を中央、垂直位置を中央
cell.alignment = Alignment(horizontal="center", vertical="center")

# 罫線の指定
cell.border = Border(
    top=Side(style="thin", color="000000"),  # 上
    right=Side(style="thin", color="000000"),  # 右
    bottom=Side(style="thin", color="000000"),  # 下
    left=Side(style="thin", color="000000"),  # 左
)

# フォントの指定
# 文字サイズ、太文字、イタリック、文字色白
cell.font = Font(size=14, bold=True, italic=True, color="FFFFFF")

# 背景色の指定
# 塗りつぶし、赤色
cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")

# 保存
book.save("style.xlsx")
