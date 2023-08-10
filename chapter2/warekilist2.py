import openpyxl as excel

# 西暦和暦対応表を作成(Excel計算機能利用版)

# 新規ワークブックを作ってシートを得る
book = excel.Workbook()
sheet = book.active
# シートのヘッダ部分に説明を入れる
sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

# 100年分の西暦和暦の対応表を作る
start_y = 1930
for i in range(100):
    # 西暦と和暦を計算
    sei = start_y + i
    wa = f'=TEXT("{sei}/1/1","ggge年")'
    # シートに設定
    sheet.cell(2 + i, 1, f"{sei}年")
    sheet.cell(2 + i, 2, wa)
    print(f"{sei}={wa}")

# ファイルを保存
book.save("wareki2.xlsx")
