import openpyxl as excel

book = excel.load_workbook("uriage.xlsx", data_only=True)  # data_onlyで計算済の値を持ってくる
sheet = book.active

# iter_rowsを使って全データを取り出す
for row in sheet.iter_rows(min_row=3):
    values = [cell.value for cell in row]
    if values[0] is None:
        break
    print(values)
