import openpyxl as excel

# 顧客一覧のブックを開く
book = excel.load_workbook("all-customer.xlsx")
# 名簿のシートを選択
sheet = book["名簿"]

# 顧客一覧を確認してシートに分ける
for row in sheet.iter_rows(min_row=3):
    cells = [v.value for v in row]
    if cells[0] is None:
        break
    # 各列のデータを変数に代入
    (name, area, plan) = cells
    # コピー先のシート名を決める
    sname = f"{plan}プラン"
    # 該当するシートがあるか
    if sname not in book.sheetnames:
        to_sheet = book.create_sheet(title=sname)
        to_sheet.append(["名前", "住所", "プラン"])
    else:
        to_sheet = book[sname]
    # 該当シートに顧客情報を追記
    to_sheet.append(cells)

# ファイルに書き込む
book.save("split_sheet.xlsx")
