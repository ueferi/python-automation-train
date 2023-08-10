import openpyxl as excel
import random

# シートを100枚作り、そのうちの1枚を当たりとする
# 当たりシートを捜すゲーム

# 当たりシートの番号を決める
atari = random.randint(1, 100)

book = excel.Workbook()
book.active["B2"]

# 繰り返し100回シートを作成
for i in range(1, 101):
    # 新規シートを作成
    sname = f"{i}番"
    sheet = book.create_sheet(title=sname)
    # シートに書き込む単語を決定
    word = "ハズレ"
    if i == atari:
        word = "当たり"
    # インパクトがあるように画面をyで埋める
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y + 1, x + 1)
            c.value = word

book.save("game100.xlsx")
print("ok, atari=", atari)  # 当たりを表示
