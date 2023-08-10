import openpyxl as excel
import datetime

# 小学校の入学年度を計算する
# 新規ワークブックを作りワークシートを得る
book = excel.Workbook()
sheet = book.active

# 開始年を設定
base_year = datetime.date.today().year - 30

# 連続でセルに値を設定する
for i in range(50):
    # 設定する値を計算
    year = base_year + i
    s1 = year - 7  # 4/2以降に生まれた人
    s2 = year - 6  # 早生まれの人
    sf = f"{s1}年4/2〜{s2}年4/1に生まれた人"
    # セルに値を設定
    sheet.cell(i + 1, 1, value=f"{year}年度")
    sheet.cell(i + 1, 2, value=sf)

# ファイルを保存
book.save("nyugaku_list.xlsx")
